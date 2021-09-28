import math

from sklearn import svm
from ezc3d import c3d
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import random
import pandas as pd
import openpyxl
import Datafiting
import xlrd
import xlwt
import xlutils
import os
import copy

def produce(str,kk,left,right):
    a = [([0] * 3) for i in range(kk)]
    t1 = [[0] for i in range(kk)]
    t2 = [[0] for i in range(kk)]
    t3 = [[0] for i in range(kk)]
    L = np.empty([1, kk * 3], dtype=float)
    k = used_label[str]
    p=random.randint(left,right)
    #print("p=",p)
    for i in range(kk):
        frame = point_data[:, :, i]
        x = frame[0, :]
        y = frame[1, :]
        z = frame[2, :]
        # print("x[1]=",x[1],"i=",i)
        if (math.isnan(x[k])):  # 这个函数来判断数值是否是NaN
            x[k] = 0;
            y[k] = 0;
            z[k] = 0
        a[i] = [x[k], y[k], z[k]]
        t1[i] = x[k]+p
        t2[i] = y[k]
        t3[i] = z[k]
        #print("order=",i," ",x[k],y[k],z[k])
        L[0, i * 3] = x[k]+p
        L[0, i * 3 + 1] = y[k]
        L[0, i * 3 + 2] = z[k]
    #print(L.shape)
    #print(L[0,0])
    return L

def locate(str,kk):
    a = [([0] * 3) for i in range(kk)]
    t1 = [[0] for i in range(kk)]
    t2 = [[0] for i in range(kk)]
    t3 = [[0] for i in range(kk)]
    L = np.empty([1, kk * 3], dtype=float)
    k = used_label[str]
    print("k=",k)
    start=0
    flag=True
    maxinum=0
    minimum=99999
    for i in range(kk):
        frame = point_data[:, :, i]
        x = frame[0, :]
        y = frame[1, :]
        z = frame[2, :]
        # print("x[1]=",x[1],"i=",i)
        if (math.isnan(x[k])):  # 这个函数来判断数值是否是NaN
            x[k] = 0
            y[k] = 0
            z[k] = 0
        else:
            if(flag):
                start=i
                flag=False
                print("start=",start)
        a[i] = [x[k], y[k], z[k]]
        t1[i] = x[k]
        t2[i] = y[k]
        t3[i] = z[k]
        if t3[i]<minimum:
            minimum=t3[i]
        if t3[i]>maxinum:
            maxinum=t3[i]
        #print("order=",i," ",x[k],y[k],z[k],t3[i])
        L[0, i * 3] = x[k]
        L[0, i * 3 + 1] = y[k]
        L[0, i * 3 + 2] = z[k]
    maxima_values = [0 for i in range(kk)]
    m=-1
    maxima_values_order=-1
    minima_values = [0 for i in range(kk)]
    minima_values_order=-1
    dividing_line=(maxinum+minimum)/2
    print("dividing_line=",dividing_line)
    boundary=[]
    for i in range(start+1,kk-1):
        if (t3[i-1]<t3[i]>t3[i+1]) and (t3[i]>dividing_line):
            maxima_values_order+=1
            maxima_values[maxima_values_order]=i
        if (t3[i-1]>t3[i]<t3[i+1]) and (t3[i]<dividing_line) and (t3[i]!=0): #保证极小值点不会是缺失点
            minima_values_order+=1
            minima_values[minima_values_order]=i
            print(t3[i-1],t3[i],t3[i+1],i)
    print(maxima_values)
    print(minima_values)
    for i in range(0,maxima_values_order):
        left=maxima_values[i]
        right=maxima_values[i+1]
        print("left=%d,right=%d"%(left,right))
        for j in range(minima_values_order+1):
            #print("minima_values[j]=",minima_values[j])
            if (minima_values[j]>left):
                left2=minima_values[j]
                left2_order=j
                #print("left2=",left2,"left=",left)
                break
        #print("left2=%d" % (left2))
        l_min_r=9999
        l_min_r_order=-1
        for j in range(left2_order,minima_values_order+1):
            print("j=",j)
            if(minima_values[j]<right):
                if(l_min_r>t3[minima_values[j]]):
                    l_min_r=t3[minima_values[j]]
                    l_min_r_order=minima_values[j]
        print("l_min_r_order=",l_min_r_order)
        if(l_min_r_order!=-1):
            boundary.append(l_min_r_order)
        print("分界点索引：%d ；z坐标：%f" % (l_min_r_order,t3[l_min_r_order]))

    print(boundary)
    #print("t3=",t3[721])
    # print(L.shape)
    # print(L[0,0])
    return boundary

if __name__ == '__main__':
    #path =r"E:\PythonProjects\pythonproject1\gujian20210519\Cal 08.c3d"
    #path="1Cal 25.c3d"
    #path="01_01.c3d"
    #path =r"E:\PythonProjects\pythonproject1\dengzhongmei\dzm20201202\Cal 03.c3d"
    path=r"E:\data\data\dzm\Cal 07.c3d"
    #path=r"E:\data\data\tx\Cal 13.c3d"
    str1 = 'dzmCal 07.csv'
    c=c3d(path)
    print(c['parameters']['POINT'])
    #print(len(c['parameters']['POINT']['LABELS']))
    used_label = {"LASI":0, "RASI":0, "LPSI":0,"RPSI":0 ,"LTHI":0,"RTHI":0,"LKNE":0,"RKNE":0,"LTIB":0,"RTIB":0,"LANK":0,"RANK":0,"LHEE":0,"RHEE":0,"LTOE":0,"RTOE":0}
    label_keys = used_label.keys()
    labels = c['parameters']['POINT']['LABELS']['value']
    point_data = c['data']['points']
    ordernum=0
    print(type(labels))
    print(labels)
    for x in labels:
        #print(type(labels))
        #print(ordernum)
        for label in label_keys:
            if (label==x):
                #print(x)
                #print("used_label.before=",used_label[x])
                used_label[x]=ordernum
                #print("used_label.after=",used_label[x])
        ordernum+=1
    #print("ordernum=",ordernum)
    length=len(point_data)
    print("length=",length)
    kk=0
    for x in point_data[1,1,:]:
        kk+=1
    print("数据帧总数=",kk)
    a=[([0] * 3)for i in range(kk)]
    t1=[[0] for i in range(kk)]
    t2=[[0] for i in range(kk)]
    t3=[[0] for i in range(kk)]
    L = np.empty([117, kk * 3], dtype=float)
    k=used_label['RKNE']
    for i in range(kk):
        frame = point_data[:, :, i]
        x = frame[0, :]
        y = frame[1, :]
        z = frame[2, :]
        #print("x[1]=",x[1],"i=",i)
        if (math.isnan(x[k])):    #这个函数来判断数值是否是NaN
            x[k]=0; y[k]=0; z[k]=0
        a[i]=[x[k],y[k],z[k]]
        t1[i] = x[k]
        t2[i] = y[k]
        t3[i] = z[k]
        #print("order=",x[k],y[k],z[k])
        L[116,i*3]=x[k]
        L[116,i*3+1]=y[k]
        L[116,i*3+2]=z[k]
    print("type(L)=",L.shape)
    #print(L[0,1])

    print(random.randint(30,50))
    LL=locate("RHEE", kk)
    LL2=locate("LHEE",kk)
    print("boundary=",LL)
    print("boundary=",LL2)
    LLL=np.array(LL)
    k = used_label['RHEE']
    datat1 = [[0] for i in range(kk)]
    datat2 = [[0] for i in range(kk)]
    datat3 = [[0] for i in range(kk)]
    for i in range(kk):
        frame = point_data[:, :, i]
        x = frame[0, :]
        y = frame[1, :]
        z = frame[2, :]
        # print("x[1]=",x[1],"i=",i)
        if (math.isnan(x[k])):  # 这个函数来判断数值是否是NaN
            x[k] = 0;
            y[k] = 0;
            z[k] = 0
        datat1[i] = x[k]
        datat2[i] = y[k]
        datat3[i] = z[k]
    '''
    for i in range(300,304):
        print("datat3[%d]=%f" % (i,datat3[i]))
        datat3[i]=0
    '''

    #print(LLL.shape)

    writer = pd.ExcelWriter("E:\PythonProjects\pythonproject1\gait_analysis\data2.xlsx",engine='openpyxl')
    for i in range(LLL.shape[0]-1):
        az = []
        for j in range(LLL[i],LLL[i+1]):
            if(datat3[j]==0):
                #ttt=datafit(datat3,j-10,j+20)
                ttt=Datafiting.datafit(datat3,LLL[i],LLL[i+1])
                for k in range(LLL[i],LLL[i+1]):
                    #print((ttt[k-LLL[i]+1]),datat3[k])
                    datat3[k]=ttt[k-LLL[i]]
                break
        for j in range(LLL[i], LLL[i + 1]):
            az.append(datat3[j])
        df=pd.DataFrame(az,index=None)
        df = pd.DataFrame(df.values.T, index=df.columns, columns=df.index)
        #print("df=", df)

        df.to_excel(writer, index=False, sheet_name="data2")
        writer.save() # to_excel方法

        filename=r"E:\PythonProjects\pythonproject1\gait_analysis\data2.xlsx"
    print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    num=0
    for i in range(LLL.shape[0] - 1):
        #xyz=[[0]*cols for xyzi in range(rows)]
        xuhao = 0
        for name in used_label:
            xuhao += 1
            num+=3
            print("name=", name, "索引=", used_label[name], "序号=", xuhao)
            print(" 第%d个步态周期"%(i+1))
            k = used_label[name]
            xt1 = [[0] for i in range(kk)]
            yt2 = [[0] for i in range(kk)]
            zt3 = [[0] for i in range(kk)]
            for i2 in range(kk):
                frame = point_data[:, :, i2]
                x = frame[0, :]
                y = frame[1, :]
                z = frame[2, :]
                # print("x[1]=",x[1],"i=",i)
                if (math.isnan(x[k])):  # 这个函数来判断数值是否是NaN
                    x[k] = 0;
                    y[k] = 0;
                    z[k] = 0
                xt1[i2] = x[k]
                yt2[i2] = y[k]
                zt3[i2] = z[k]
            for j in range(LLL[i], LLL[i + 1]):
                if (zt3[j] == 0):
                    ttt = Datafiting.datafit(zt3, LLL[i], LLL[i + 1])
                    for ii in range(LLL[i], LLL[i + 1]):
                        # print((ttt[k-LLL[i]+1]),datat3[k])
                        zt3[ii] = ttt[ii - LLL[i]]
                    break
            for j in range(LLL[i], LLL[i + 1]):
                if (xt1[j] == 0):
                    ttt = Datafiting.datafit(xt1, LLL[i], LLL[i + 1])
                    for ii in range(LLL[i], LLL[i + 1]):
                        # print((ttt[k-LLL[i]+1]),datat3[k])
                        xt1[ii] = ttt[ii - LLL[i]]
                    break
            for j in range(LLL[i], LLL[i + 1]):
                if (yt2[j] == 0):
                    ttt = Datafiting.datafit(yt2, LLL[i], LLL[i + 1])
                    for ii in range(LLL[i], LLL[i + 1]):
                        # print((ttt[k-LLL[i]+1]),datat3[k])
                        yt2[ii] = ttt[ii - LLL[i]]
                    break
            xyzx=[]
            xyzy=[]
            xyzz=[]
            for j in range(LLL[i], LLL[i + 1]):
                xyzx.append(xt1[j])
                xyzy.append(yt2[j])
                xyzz.append(zt3[j])
            xyzx=np.array(xyzx)

            dataframe=pd.DataFrame(xyzx,index=None)
            #print("dataframe.shape=",dataframe.shape)
            #print("dataframe.type=",type(dataframe))
            #print("dataframe[4]=",dataframe[4])
            dataframe=pd.DataFrame(dataframe.values.T, index=dataframe.columns, columns=dataframe.index)
            dataframe.to_csv(str1, mode='a', header=False, index=False)
            dataframe = pd.DataFrame(xyzy, index=None)
            #print("dataframe.shape=", dataframe.shape)
            #print("dataframe.type=", type(dataframe))
            # print("dataframe[4]=",dataframe[4])
            dataframe = pd.DataFrame(dataframe.values.T, index=dataframe.columns, columns=dataframe.index)
            dataframe.to_csv(str1, mode='a', header=False, index=False)
            dataframe = pd.DataFrame(xyzz, index=None)
            #print("dataframe.shape=", dataframe.shape)
            #print("dataframe.type=", type(dataframe))
            # print("dataframe[4]=",dataframe[4])
            dataframe = pd.DataFrame(dataframe.values.T, index=dataframe.columns, columns=dataframe.index)
            dataframe.to_csv(str1, mode='a', header=False, index=False)
        #writer.book=load_workbook("E:\PythonProjects\pythonproject1\gait_analysis\data2.xlsx")
        print("总数=",num)
'''     workbook=xlrd.open_workbook('data.xls',formatting_info=True)
        sheet = workbook.sheet_by_index(0)
        print(sheet)
        rowNum=sheet.nrows
        colNum=sheet.ncols
        newbook=copy.copy(workbook)
        print("newbook=",newbook)
        wb=openpyxl.Workbook()
        wb=openpyxl.load_workbook('data2.xlsx')
        ws=wb.copy_worksheet("data2")
        newsheet=newbook.get_sheet(0)
        newsheet.write(rowNum,0,df)
        ws.append(df)
        newbook.save('data.xls')'''





